from typing import List, Optional, Dict, Any
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle,
    GradientFill,
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
from io import BytesIO

from .models import PerDcomp


class PerDcompExcelExporter:
    """
    Enhanced service for exporting PER/DCOMP data to Excel format with advanced features:
    - Performance optimization for large datasets
    - Professional formatting with conditional formatting
    - Interactive tables and charts
    - Memory-efficient processing
    """

    def __init__(self, chunk_size: int = 1000):
        self.workbook = None
        self.worksheet = None
        self.chunk_size = chunk_size
        self._setup_styles()

    def _setup_styles(self):
        """Pre-define reusable styles for better performance."""
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color="FFFFFF", size=12)
        self.header_style.fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.currency_style = NamedStyle(name="currency")
        self.currency_style.number_format = "R$ #,##0.00"
        self.currency_style.alignment = Alignment(horizontal="right")

        self.date_style = NamedStyle(name="date")
        self.date_style.number_format = "DD/MM/YYYY"
        self.date_style.alignment = Alignment(horizontal="center")

    def export_to_excel(
        self,
        queryset,
        client_cnpj: Optional[str] = None,
        applied_filters: Optional[dict] = None,
        optimize_for_size: bool = True,
    ) -> HttpResponse:
        """
        Export PER/DCOMP queryset to Excel file with advanced formatting.

        Args:
            queryset: Filtered PerDcomp queryset
            client_cnpj: Optional CNPJ for filename
            applied_filters: Dict of filters applied for metadata
            optimize_for_size: Whether to optimize for file size vs features

        Returns:
            HttpResponse with Excel file
        """
        self.workbook = Workbook()

        # Register custom styles
        if self.header_style.name not in self.workbook.named_styles:
            self.workbook.add_named_style(self.header_style)
        if self.currency_style.name not in self.workbook.named_styles:
            self.workbook.add_named_style(self.currency_style)
        if self.date_style.name not in self.workbook.named_styles:
            self.workbook.add_named_style(self.date_style)

        # Create worksheets with enhanced features
        self._create_data_sheet(queryset, optimize_for_size)
        self._create_summary_sheet(queryset, client_cnpj, applied_filters)
        self._create_status_analysis_sheet(queryset)

        # Generate filename with metadata
        filename = self._generate_enhanced_filename(client_cnpj, applied_filters)

        # Prepare optimized response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"

        # Save workbook with memory optimization
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())
        buffer.close()

        return response

    def _create_data_sheet(self, queryset, optimize_for_size=False):
        """Create enhanced main data sheet with Excel tables and conditional formatting."""
        self.worksheet = self.workbook.active
        self.worksheet.title = "PER-DCOMPs"

        # Enhanced headers with better descriptions
        headers = [
            "Número PER/DCOMP",
            "Número",
            "Protocolo Processo",
            "CNPJ",
            "Razão Social",
            "Status",
            "Valor Pedido (R$)",
            "Valor Compensado (R$)",
            "Valor Recebido (R$)",
            "Valor Saldo (R$)",
            "Valor SELIC (R$)",
            "Tributo",
            "Competência",
            "Data Transmissão",
            "Data Vencimento",
            "Data Competência",
            "Dias até Vencimento",
            "Criado em",
        ]

        # Add headers with enhanced styling
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            if (
                hasattr(self, "header_style")
                and self.header_style.name in self.workbook.named_styles
            ):
                cell.style = self.header_style
            else:
                # Fallback styling
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="2F5597", end_color="2F5597", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Process data with performance optimization
        total_count = queryset.count()
        row_num = 2

        # Pre-calculate values for performance
        current_date = datetime.now().date()

        # Process in chunks if we have many records
        chunk_size = (
            getattr(self, "chunk_size", 1000) if total_count > 1000 else total_count
        )

        for chunk_start in range(0, total_count, chunk_size):
            chunk = queryset[chunk_start : chunk_start + chunk_size]

            for perdcomp in chunk:
                # Calculate days until expiration
                days_until_expiration = ""
                if perdcomp.data_vencimento:
                    delta = perdcomp.data_vencimento - current_date
                    days_until_expiration = delta.days

                data = [
                    perdcomp.numero_perdcomp,
                    perdcomp.numero or "",
                    perdcomp.processo_protocolo or "",
                    perdcomp.cnpj,
                    (
                        getattr(perdcomp.client, "razao_social", "")
                        if perdcomp.client
                        else ""
                    ),
                    perdcomp.get_status_display(),
                    self._format_currency(perdcomp.valor_pedido),
                    self._format_currency(perdcomp.valor_compensado),
                    self._format_currency(perdcomp.valor_recebido),
                    self._format_currency(perdcomp.valor_saldo),
                    self._format_currency(perdcomp.valor_selic),
                    perdcomp.tributo_pedido,
                    perdcomp.competencia or "",
                    perdcomp.data_transmissao,
                    perdcomp.data_vencimento,
                    perdcomp.data_competencia,
                    days_until_expiration,
                    perdcomp.created_at.date() if perdcomp.created_at else "",
                ]

                for col, value in enumerate(data, 1):
                    cell = self.worksheet.cell(row=row_num, column=col)
                    cell.value = value

                    # Apply enhanced formatting
                    if col in [7, 8, 9, 10, 11]:  # Currency columns
                        if isinstance(value, (int, float, Decimal)):
                            if (
                                hasattr(self, "currency_style")
                                and self.currency_style.name
                                in self.workbook.named_styles
                            ):
                                cell.style = self.currency_style
                            else:
                                cell.number_format = "R$ #,##0.00"
                                cell.alignment = Alignment(horizontal="right")
                    elif col in [14, 15, 16, 18]:  # Date columns
                        if value:
                            if (
                                hasattr(self, "date_style")
                                and self.date_style.name in self.workbook.named_styles
                            ):
                                cell.style = self.date_style
                            else:
                                cell.number_format = "DD/MM/YYYY"
                                cell.alignment = Alignment(horizontal="center")
                    elif col == 17:  # Days until expiration
                        if isinstance(value, int):
                            cell.alignment = Alignment(horizontal="center")
                            # Color code based on urgency (only if not optimizing for size)
                            if not optimize_for_size:
                                if value < 0:
                                    cell.fill = PatternFill(
                                        start_color="FFCCCB",
                                        end_color="FFCCCB",
                                        fill_type="solid",
                                    )  # Light red for overdue
                                elif value < 30:
                                    cell.fill = PatternFill(
                                        start_color="FFFFCC",
                                        end_color="FFFFCC",
                                        fill_type="solid",
                                    )  # Light yellow for urgent
                                elif value < 90:
                                    cell.fill = PatternFill(
                                        start_color="E6F3FF",
                                        end_color="E6F3FF",
                                        fill_type="solid",
                                    )  # Light blue for warning

                    # Add borders for better readability (only if not optimizing for size)
                    if not optimize_for_size:
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                row_num += 1

        # Create Excel table and conditional formatting (only if not optimizing for size)
        if not optimize_for_size and total_count > 0:
            # Create Excel table for better interaction
            try:
                from openpyxl.worksheet.table import Table, TableStyleInfo

                table_range = f"A1:{get_column_letter(len(headers))}{row_num-1}"
                table = Table(displayName="PerDcompsTable", ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                self.worksheet.add_table(table)
            except Exception:
                pass  # Skip table creation if it fails

            # Add conditional formatting
            try:
                self._add_conditional_formatting()
            except Exception:
                pass  # Skip conditional formatting if it fails

        # Auto-adjust column widths
        if hasattr(self, "_adjust_column_widths_enhanced"):
            self._adjust_column_widths_enhanced()
        else:
            self._adjust_column_widths()

    def _create_summary_sheet(self, queryset, client_cnpj, applied_filters):
        """Create enhanced summary statistics sheet with professional formatting."""
        summary_sheet = self.workbook.create_sheet(title="Resumo")

        # Company header section with enhanced styling
        summary_sheet["A1"] = "RELATÓRIO EXECUTIVO - PER/DCOMPs"
        summary_sheet["A1"].font = Font(bold=True, size=18, color="FFFFFF")
        summary_sheet["A1"].fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        summary_sheet["A1"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        summary_sheet.merge_cells("A1:F1")
        summary_sheet.row_dimensions[1].height = 30

        # Get client information for better context
        client_info = self._get_client_info(client_cnpj) if client_cnpj else {}

        # Client and export information section
        row = 3
        section_color = "E7E6E6"

        # Client information
        summary_sheet[f"A{row}"] = "INFORMAÇÕES DO CLIENTE"
        summary_sheet[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        summary_sheet[f"A{row}"].fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        summary_sheet.merge_cells(f"A{row}:F{row}")
        row += 1

        # Client details with better formatting
        if client_info:
            details = [
                ("Razão Social:", client_info.get("razao_social", "N/A")),
                ("CNPJ:", client_cnpj),
                ("Nome Fantasia:", client_info.get("nome_fantasia", "N/A")),
                ("Situação:", client_info.get("status", "N/A")),
            ]
        else:
            details = [("CNPJ:", client_cnpj)]

        for i, (label, value) in enumerate(details):
            col_label = chr(65 + (i % 2) * 3)  # A or D
            col_value = chr(66 + (i % 2) * 3)  # B or E

            if i % 2 == 0 and i > 0:
                row += 1

            summary_sheet[f"{col_label}{row}"] = label
            summary_sheet[f"{col_label}{row}"].font = Font(bold=True)
            summary_sheet[f"{col_value}{row}"] = str(value)

        row += 2

        # Export metadata section
        summary_sheet[f"A{row}"] = "INFORMAÇÕES DA EXPORTAÇÃO"
        summary_sheet[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        summary_sheet[f"A{row}"].fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        summary_sheet.merge_cells(f"A{row}:F{row}")
        row += 1

        export_info = [
            ("Data/Hora da Exportação:", datetime.now().strftime("%d/%m/%Y às %H:%M")),
            ("Total de Registros:", f"{queryset.count():,}".replace(",", ".")),
        ]

        # Add filter information if any (excluding redundant client_cnpj)
        if applied_filters:
            clean_filters = {
                k: v for k, v in applied_filters.items() if k != "client_cnpj" and v
            }
            if clean_filters:
                filter_text = ", ".join(
                    [
                        f"{k.replace('_', ' ').title()}: {v}"
                        for k, v in clean_filters.items()
                    ]
                )
                export_info.append(("Filtros Aplicados:", filter_text))

        for i, (label, value) in enumerate(export_info):
            summary_sheet[f"A{row + i}"] = label
            summary_sheet[f"A{row + i}"].font = Font(bold=True)
            summary_sheet[f"B{row + i}"] = str(value)

        row += len(export_info) + 2

        # Financial overview section
        summary_sheet[f"A{row}"] = "RESUMO FINANCEIRO"
        summary_sheet[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        summary_sheet[f"A{row}"].fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        summary_sheet.merge_cells(f"A{row}:F{row}")
        row += 1

        # Calculate enhanced financial statistics
        total_value = self._calculate_total_value(queryset)
        avg_value = (
            total_value / queryset.count() if queryset.count() > 0 else Decimal("0")
        )

        # Calculate other financial metrics
        compensated_total = Decimal("0")
        received_total = Decimal("0")
        balance_total = Decimal("0")

        for perdcomp in queryset.only(
            "valor_compensado", "valor_recebido", "valor_saldo"
        ):
            try:
                if perdcomp.valor_compensado:
                    compensated_total += Decimal(
                        str(perdcomp.valor_compensado).replace(",", ".")
                    )
                if perdcomp.valor_recebido:
                    received_total += Decimal(
                        str(perdcomp.valor_recebido).replace(",", ".")
                    )
                if perdcomp.valor_saldo:
                    balance_total += Decimal(
                        str(perdcomp.valor_saldo).replace(",", ".")
                    )
            except (ValueError, InvalidOperation):
                continue

        financial_stats = [
            (
                "Valor Total Pedido:",
                f"R$ {total_value:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            ),
            (
                "Valor Médio por PER/DCOMP:",
                f"R$ {avg_value:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            ),
            (
                "Valor Total Compensado:",
                f"R$ {compensated_total:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            ),
            (
                "Valor Total Recebido:",
                f"R$ {received_total:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            ),
            (
                "Valor Total Saldo:",
                f"R$ {balance_total:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            ),
        ]

        for i, (label, value) in enumerate(financial_stats):
            col_label = chr(65 + (i % 2) * 3)  # A or D
            col_value = chr(66 + (i % 2) * 3)  # B or E

            if i % 2 == 0 and i > 0:
                row += 1

            summary_sheet[f"{col_label}{row}"] = label
            summary_sheet[f"{col_label}{row}"].font = Font(bold=True)
            summary_sheet[f"{col_value}{row}"] = value
            summary_sheet[f"{col_value}{row}"].font = Font(bold=True)

        row += 2

        # Status analysis section
        summary_sheet[f"A{row}"] = "ANÁLISE DE STATUS"
        summary_sheet[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        summary_sheet[f"A{row}"].fill = PatternFill(
            start_color="C65911", end_color="C65911", fill_type="solid"
        )
        summary_sheet.merge_cells(f"A{row}:F{row}")
        row += 1

        # Enhanced status breakdown
        total_count = queryset.count()
        status_analysis = {}

        for status_choice in PerDcomp.Status.choices:
            status_code, status_display = status_choice
            count = queryset.filter(status=status_code).count()
            if count > 0:
                percentage = (count / total_count * 100) if total_count > 0 else 0
                status_analysis[status_display] = {
                    "count": count,
                    "percentage": percentage,
                }

        # Create table headers
        headers = ["Status", "Quantidade", "Percentual"]
        for col, header in enumerate(headers):
            cell = summary_sheet.cell(row=row, column=col + 1)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        row += 1

        # Add status data
        for status, data in status_analysis.items():
            for col, value in enumerate(
                [status, data["count"], f"{data['percentage']:.1f}%"]
            ):
                cell = summary_sheet.cell(row=row, column=col + 1)
                cell.value = value
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if col > 0:  # Align numbers to center
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        row += 1

        # Top tributos section
        tributo_counts = {}
        for perdcomp in queryset.only("tributo_pedido"):
            tributo = perdcomp.tributo_pedido
            if tributo:
                tributo_counts[tributo] = tributo_counts.get(tributo, 0) + 1

        if tributo_counts:
            summary_sheet[f"A{row}"] = "TOP TRIBUTOS"
            summary_sheet[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
            summary_sheet[f"A{row}"].fill = PatternFill(
                start_color="7030A0", end_color="7030A0", fill_type="solid"
            )
            summary_sheet.merge_cells(f"A{row}:F{row}")
            row += 1

            # Sort and get top 5 tributos
            top_tributos = sorted(
                tributo_counts.items(), key=lambda x: x[1], reverse=True
            )[:5]

            for tributo, count in top_tributos:
                percentage = (count / total_count * 100) if total_count > 0 else 0
                summary_sheet[f"A{row}"] = tributo
                summary_sheet[f"A{row}"].font = Font(bold=True)
                summary_sheet[f"B{row}"] = f"{count} ({percentage:.1f}%)"
                row += 1

        # Enhanced column formatting
        column_widths = {"A": 25, "B": 20, "C": 15, "D": 25, "E": 20, "F": 15}

        for col, width in column_widths.items():
            summary_sheet.column_dimensions[col].width = width

        # Add subtle background to alternating rows for better readability
        for row_num in range(1, row + 1):
            if row_num % 2 == 0:
                for col_num in range(1, 7):
                    cell = summary_sheet.cell(row=row_num, column=col_num)
                    if (
                        not cell.fill.start_color.rgb
                        or cell.fill.start_color.rgb == "00000000"
                    ):
                        cell.fill = PatternFill(
                            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                        )

    def _get_client_info(self, client_cnpj):
        """Get client information for better context in summary."""
        try:
            from apps.clients.models import Client

            client = Client.objects.filter(
                cnpj=client_cnpj, deleted_at__isnull=True
            ).first()
            if client:
                return {
                    "razao_social": getattr(client, "razao_social", ""),
                    "nome_fantasia": getattr(client, "nome_fantasia", ""),
                    "status": getattr(client, "status", ""),
                }
        except Exception:
            pass
        return {}

    def _format_currency(self, value) -> Optional[float]:
        """Convert string currency values to float for Excel formatting."""
        if not value:
            return None
        try:
            value_str = str(value).replace(",", ".")
            return float(Decimal(value_str))
        except (ValueError, InvalidOperation):
            return None

    def _calculate_total_value(self, queryset) -> Decimal:
        """Calculate total value from queryset."""
        total = Decimal("0.00")
        for perdcomp in queryset.only("valor_pedido"):
            try:
                valor_str = str(perdcomp.valor_pedido or "0.00").replace(",", ".")
                valor_decimal = Decimal(valor_str)
                total += valor_decimal
            except (ValueError, InvalidOperation):
                continue
        return total

    def _adjust_column_widths(self):
        """Auto-adjust column widths based on content."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set minimum width and add some padding
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = max(
                adjusted_width, 12
            )

    def _generate_filename(self, client_cnpj: Optional[str]) -> str:
        """Generate appropriate filename for the export."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if client_cnpj:
            return f"perdcomps_export_{client_cnpj}_{timestamp}.xlsx"
        else:
            return f"perdcomps_export_{timestamp}.xlsx"

    def _adjust_column_widths_enhanced(self):
        """Enhanced auto-adjust column widths with performance optimization."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Sample only first 100 rows for performance
            for cell in list(column)[:100]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set optimal width with limits
            if max_length < 8:
                adjusted_width = 12
            elif max_length > 40:
                adjusted_width = 40
            else:
                adjusted_width = max_length + 3

            self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_conditional_formatting(self):
        """Add conditional formatting for better visual analysis."""
        if self.worksheet.max_row <= 1:
            return

        # Color scale for currency values (column G - Valor Pedido)
        currency_range = f"G2:K{self.worksheet.max_row}"
        color_scale_rule = ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB9C",
            end_type="max",
            end_color="FFC000",
        )
        self.worksheet.conditional_formatting.add(currency_range, color_scale_rule)

        # Data bars for "Days until expiration" (column Q)
        if self.worksheet.max_column >= 17:
            days_range = f"Q2:Q{self.worksheet.max_row}"
            data_bar_rule = DataBarRule(
                start_type="num",
                start_value=-30,
                end_type="num",
                end_value=365,
                color="6FA8DC",
            )
            self.worksheet.conditional_formatting.add(days_range, data_bar_rule)

    def _create_status_analysis_sheet(self, queryset):
        """Create executive-friendly status analysis sheet with clear tables and metrics."""
        analysis_sheet = self.workbook.create_sheet(title="Análise de Status")

        # Professional sheet settings
        analysis_sheet.sheet_view.showGridLines = False

        # ===========================================
        # SECTION 1: EXECUTIVE HIGHLIGHTS (TOP)
        # ===========================================

        # Calculate key metrics first
        total_count = queryset.count()
        total_requested = self._calculate_total_value(queryset)
        total_received = self._calculate_field_total(queryset, "valor_recebido")
        total_balance = self._calculate_field_total(queryset, "valor_saldo")

        # Main title
        analysis_sheet["A1"] = "ANÁLISE EXECUTIVA DE STATUS"
        analysis_sheet["A1"].font = Font(bold=True, size=18, color="FFFFFF")
        analysis_sheet["A1"].fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        analysis_sheet["A1"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        analysis_sheet.merge_cells("A1:F1")
        analysis_sheet.row_dimensions[1].height = 35

        # Executive summary cards (2x2 grid)
        self._create_executive_cards(
            analysis_sheet,
            {
                "total_count": total_count,
                "total_requested": total_requested,
                "total_received": total_received,
                "total_balance": total_balance,
            },
        )

        # ===========================================
        # SECTION 2: DETAILED STATUS TABLE
        # ===========================================

        # Table starts at row 8 (after cards and spacing)
        table_start_row = 8

        # Section title
        analysis_sheet[f"A{table_start_row}"] = "DETALHAMENTO POR STATUS"
        analysis_sheet[f"A{table_start_row}"].font = Font(
            bold=True, size=14, color="2F5597"
        )
        analysis_sheet[f"A{table_start_row}"].alignment = Alignment(horizontal="left")
        analysis_sheet.merge_cells(f"A{table_start_row}:F{table_start_row}")
        analysis_sheet.row_dimensions[table_start_row].height = 25

        # Table headers
        headers_row = table_start_row + 2
        headers = [
            "Status",
            "Quantidade",
            "% do Total",
            "Valor Pedido (R$)",
            "Valor Recebido (R$)",
            "Saldo (R$)",
        ]

        # Create professional table headers
        for col_num, header in enumerate(headers, 1):
            cell = analysis_sheet.cell(headers_row, col_num, header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="2F5597", end_color="2F5597", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            )
        analysis_sheet.row_dimensions[headers_row].height = 25

        # Calculate status data
        status_data = self._calculate_status_metrics(queryset, total_count)

        # Sort by total requested value (descending)
        sorted_statuses = sorted(
            status_data.items(), key=lambda x: x[1]["total_requested"], reverse=True
        )

        # Add data rows
        current_row = headers_row + 1
        for status_name, metrics in sorted_statuses:
            self._create_status_data_row(
                analysis_sheet, current_row, status_name, metrics
            )
            current_row += 1

        # Add total row
        self._create_total_row(
            analysis_sheet,
            current_row,
            {
                "total_count": total_count,
                "total_requested": total_requested,
                "total_received": total_received,
                "total_balance": total_balance,
            },
        )

        # Set optimal column widths
        column_widths = {"A": 20, "B": 12, "C": 12, "D": 20, "E": 20, "F": 20}
        for col, width in column_widths.items():
            analysis_sheet.column_dimensions[col].width = width

    def _create_executive_cards(self, ws, metrics):
        """Create executive summary cards in 2x2 grid layout."""
        cards_data = [
            (
                "Total de Registros",
                f"{metrics['total_count']:,}".replace(",", "."),
                "70AD47",
            ),
            (
                "Valor Total Pedido",
                f"R$ {metrics['total_requested']:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
                "2F5597",
            ),
            (
                "Valor Total Recebido",
                f"R$ {metrics['total_received']:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
                "C65911",
            ),
            (
                "Saldo Pendente",
                f"R$ {metrics['total_balance']:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
                "7030A0",
            ),
        ]

        # Create cards in 2x2 grid (rows 3-6, columns A-F)
        card_positions = [
            ("A3", "C3", "A4", "C4"),  # Top-left
            ("D3", "F3", "D4", "F4"),  # Top-right
            ("A5", "C5", "A6", "C6"),  # Bottom-left
            ("D5", "F5", "D6", "F6"),  # Bottom-right
        ]

        for i, (label, value, color) in enumerate(cards_data):
            label_range, value_range = card_positions[i][:2], card_positions[i][2:]

            # Label cell
            ws[label_range[0]] = label
            ws[label_range[0]].font = Font(bold=True, size=10, color="FFFFFF")
            ws[label_range[0]].fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws[label_range[0]].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.merge_cells(f"{label_range[0]}:{label_range[1]}")

            # Value cell
            ws[value_range[0]] = value
            ws[value_range[0]].font = Font(bold=True, size=14, color=color)
            ws[value_range[0]].fill = PatternFill(
                start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
            )
            ws[value_range[0]].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.merge_cells(f"{value_range[0]}:{value_range[1]}")

            # Set row heights for better appearance
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 25
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 25

    def _calculate_status_metrics(self, queryset, total_count):
        """Calculate comprehensive metrics for each status."""
        status_metrics = {}

        for status_choice in PerDcomp.Status.choices:
            status_code, status_display = status_choice
            status_queryset = queryset.filter(status=status_code)
            count = status_queryset.count()

            if count > 0:
                percentage = (count / total_count) * 100 if total_count > 0 else 0
                total_requested = self._calculate_total_value(status_queryset)
                total_received = self._calculate_field_total(
                    status_queryset, "valor_recebido"
                )
                total_balance = self._calculate_field_total(
                    status_queryset, "valor_saldo"
                )

                status_metrics[status_display] = {
                    "count": count,
                    "percentage": percentage,
                    "total_requested": total_requested,
                    "total_received": total_received,
                    "total_balance": total_balance,
                }

        return status_metrics

    def _calculate_field_total(self, queryset, field_name):
        """Calculate total for a specific field in queryset."""
        total = Decimal("0.00")
        for record in queryset.only(field_name):
            try:
                value_str = str(getattr(record, field_name) or "0.00").replace(",", ".")
                value_decimal = Decimal(value_str)
                total += value_decimal
            except (ValueError, InvalidOperation):
                continue
        return total

    def _create_status_data_row(self, ws, row, status_name, metrics):
        """Create a formatted data row for status metrics."""
        # Status colors for visual distinction
        status_colors = {
            "Rascunho": "FFF2CC",  # Light yellow
            "Transmitido": "E1F5FE",  # Light blue
            "Processado": "F3E5F5",  # Light purple
            "Deferido": "E8F5E8",  # Light green
            "Indeferido": "FFEBEE",  # Light red
            "Cancelado": "F5F5F5",  # Light gray
        }

        row_data = [
            status_name,
            metrics["count"],
            metrics["percentage"] / 100,  # For percentage formatting
            float(metrics["total_requested"]),
            float(metrics["total_received"]),
            float(metrics["total_balance"]),
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row, col_num, value)

            # Apply formatting based on column
            if col_num == 1:  # Status name
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color=status_colors.get(status_name, "FFFFFF"),
                    end_color=status_colors.get(status_name, "FFFFFF"),
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_num == 2:  # Count
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=11)
            elif col_num == 3:  # Percentage
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=11)
            else:  # Currency columns
                cell.number_format = "R$ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(size=11)

            # Add borders
            cell.border = Border(
                left=Side(style="thin", color="E1E1E1"),
                right=Side(style="thin", color="E1E1E1"),
                top=Side(style="thin", color="E1E1E1"),
                bottom=Side(style="thin", color="E1E1E1"),
            )

        ws.row_dimensions[row].height = 22

    def _create_total_row(self, ws, row, totals):
        """Create professional total row at bottom of table."""
        # Add spacing row
        ws.row_dimensions[row].height = 5
        row += 1

        total_data = [
            "TOTAL",
            totals["total_count"],
            1.0,  # 100%
            float(totals["total_requested"]),
            float(totals["total_received"]),
            float(totals["total_balance"]),
        ]

        for col_num, value in enumerate(total_data, 1):
            cell = ws.cell(row, col_num, value)

            # Strong styling for total row
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="1F4E79", end_color="1F4E79", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thick", color="FFFFFF"),
                right=Side(style="thick", color="FFFFFF"),
                top=Side(style="thick", color="FFFFFF"),
                bottom=Side(style="thick", color="FFFFFF"),
            )

            # Apply formatting
            if col_num == 1:  # Label
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 2:  # Count
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 3:  # Percentage
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:  # Currency
                cell.number_format = "R$ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row].height = 30

    def _generate_enhanced_filename(
        self, client_cnpj: Optional[str], applied_filters: Optional[dict] = None
    ) -> str:
        """Generate enhanced filename with metadata."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename_parts = ["perdcomps_export"]

        if client_cnpj:
            # Clean CNPJ for filename (remove special characters)
            clean_cnpj = client_cnpj.replace(".", "").replace("/", "").replace("-", "")
            filename_parts.append(clean_cnpj)

        if applied_filters:
            if applied_filters.get("status"):
                filename_parts.append(f"status_{applied_filters['status'].lower()}")
            if applied_filters.get("search"):
                # Clean search term for filename
                search_clean = "".join(
                    c for c in applied_filters["search"] if c.isalnum()
                )[:10]
                filename_parts.append(f"search_{search_clean}")

        filename_parts.append(timestamp)

        return f"{'_'.join(filename_parts)}.xlsx"
